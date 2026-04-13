"""
Inteligência de Mercado CNP — Pipeline de coleta e padronização de dados
========================================================================
Execução automática: GitHub Actions toda segunda-feira às 08:00 UTC.
Execução manual   : python fetch_data.py

FLUXO DE DADOS
--------------
Você coloca arquivos brutos em data/raw/  →  este script processa  →  data/ (arquivos finais)

PRIORIDADE: arquivos locais em data/raw/ são SEMPRE usados se existirem.
O script só tenta baixar da internet se o arquivo local NÃO existir.
Isso significa: basta colocar o arquivo na pasta e rodar — ele não vai à internet.

ARQUIVOS QUE VOCÊ COLOCA MANUALMENTE:
  data/raw/anp_glp_YYYY.csv          → ANP GLP P13 varejo (um por ano, ex: anp_glp_2025.csv)
                                        Substituir o arquivo do ano atual a cada mês.
  data/raw/anp_gnv_YYYY.csv          → ANP GNV varejo (um por ano, ex: anp_gnv_2025.csv)
                                        Substituir o arquivo do ano atual a cada mês.
  data/raw/anp_glp_campos_YYYY.xlsx  → Preços de gás natural por campo (ex: anp_glp_campos_2025.xlsx)
                                        Atualizar quando ANP publicar nova versão.
  data/raw/ibge_pib_bruto.csv        → IBGE PIB municipal (colunas: municipio, uf, pib_total_mil)
                                        Substituir quando IBGE publicar novo ano-base.
  data/raw/cbio_bruto.csv            → CBIO3.SA exportado da B3 ou Yahoo Finance

ARQUIVOS GERADOS AUTOMATICAMENTE em data/:
  prices.csv       → preços semanais/mensais consolidados (inclui glp_campos_brl)
  ibge_mercado.csv → PIB e população por município
  cbio.csv         → histórico de preços CBIO
  last_update.txt  → data da última execução

ARQUIVO QUE VOCÊ COLOCA DIRETAMENTE (não passa por este script):
  data/empresas.csv → base CNPJ Receita Federal (coloque direto em data/, não em raw/)
"""

import yfinance as yf
import pandas as pd
import requests
import os, sys, re
from datetime import date, datetime
from io import StringIO

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
RAW_DIR  = os.path.join(DATA_DIR, "raw")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(RAW_DIR,  exist_ok=True)

TICKERS = {"brent":"BZ=F", "henry_hub":"NG=F", "diesel_us":"HO=F", "dolar":"BRL=X"}


def _arquivo_existe(caminho):
    return os.path.exists(caminho) and os.path.getsize(caminho) > 100


def _cache_valido(arquivo, dias=28):
    if not _arquivo_existe(arquivo):
        return False
    return (datetime.now().timestamp() - os.path.getmtime(arquivo)) / 86400 < dias


def fetch_yahoo():
    print("[1/6] Yahoo Finance — Brent, Henry Hub, Diesel, Dólar...")
    frames = []
    for name, ticker in TICKERS.items():
        try:
            df = yf.download(ticker, period="5y", interval="1wk",
                             auto_adjust=True, progress=False)
            if df.empty:
                print(f"  [AVISO] {name}: sem dados retornados")
                continue
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df = df[["Close"]].rename(columns={"Close": name})
            df.index = pd.to_datetime(df.index).normalize()
            frames.append(df)
            print(f"  [OK] {name}: {len(df)} semanas | último: {df.index[-1].date()}")
        except Exception as e:
            print(f"  [ERRO] {name}: {e}")
    if not frames:
        print("  [AVISO] Yahoo Finance sem dados — prices.csv será gerado apenas com ANP.")
        return pd.DataFrame()
    result = frames[0]
    for f in frames[1:]:
        result = result.join(f, how="outer")
    result.index.name = "date"
    return result.sort_index()


def _processar_anp_bruto(texto_csv, coluna_saida):
    df = pd.read_csv(StringIO(texto_csv), sep=";", decimal=",",
                     encoding="latin1", low_memory=False)
    col_data  = next((c for c in df.columns if "data" in c.lower()), None)
    col_preco = next((c for c in df.columns
                      if any(x in c.lower() for x in ["médio","medio","preco","preço"])), None)
    if not col_data or not col_preco:
        raise ValueError(f"Colunas não encontradas. Disponíveis: {list(df.columns)}")
    df[col_data]  = pd.to_datetime(df[col_data], dayfirst=True, errors="coerce")
    df[col_preco] = pd.to_numeric(
        df[col_preco].astype(str).str.replace(",", "."), errors="coerce")
    df = df.dropna(subset=[col_data, col_preco])
    serie = df.groupby(col_data)[col_preco].mean().rename(coluna_saida)
    serie.index = pd.to_datetime(serie.index)
    serie.index.name = "date"
    return serie


def fetch_anp(combinar_com=None):
    print("[2/6] ANP — GLP P13 (varejo) e GNV...")

    # URLs consolidadas (arquivo único com todo o histórico desde 2004)
    # Fonte: https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos
    ANP_URLS = {
        "glp": [
            "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/precos-revenda-e-de-distribuicao-combustiveis/serie-historica-do-levantamento-de-precos/glp.csv",
            "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/shpc/dsan/glp-automotivo-brasil-por-regiao-desde-jun2004.csv",
        ],
        "gnv": [
            "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/precos-revenda-e-de-distribuicao-combustiveis/serie-historica-do-levantamento-de-precos/gnv.csv",
            "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/shpc/dsan/gnv-automotivo-brasil-por-regiao-desde-jun2004.csv",
        ],
    }

    ano_atual = date.today().year
    resultado = {}

    for fuel, col in [("glp", "glp_p13_brl"), ("gnv", "gnv_brl")]:
        # 1) Tenta arquivos locais por ano (colocados manualmente em data/raw/)
        frames_local = []
        for ano in range(2020, ano_atual + 1):
            raw_path = os.path.join(RAW_DIR, f"anp_{fuel}_{ano}.csv")
            if _arquivo_existe(raw_path):
                try:
                    with open(raw_path, encoding="latin1") as f:
                        texto = f.read()
                    serie = _processar_anp_bruto(texto, col)
                    frames_local.append(serie)
                    print(f"  [LOCAL] ANP {fuel} {ano} — processado de data/raw/")
                except Exception as e:
                    print(f"  [AVISO] ANP {fuel} {ano} local: {e}")

        if frames_local:
            resultado[col] = pd.concat(frames_local).sort_index()
            resultado[col] = resultado[col][~resultado[col].index.duplicated(keep="last")]
            continue

        # 2) Tenta arquivo consolidado (download único com todo o histórico)
        raw_consolidado = os.path.join(RAW_DIR, f"anp_{fuel}_consolidado.csv")
        texto = None

        if _arquivo_existe(raw_consolidado):
            print(f"  [LOCAL] ANP {fuel} consolidado — usando cache em data/raw/")
            with open(raw_consolidado, encoding="latin1") as f:
                texto = f.read()
        else:
            for url in ANP_URLS[fuel]:
                try:
                    print(f"  [DOWNLOAD] ANP {fuel} consolidado — tentando {url.split('/')[-1]}...")
                    r = requests.get(url, timeout=60)
                    r.raise_for_status()
                    texto = r.content.decode("latin1")
                    with open(raw_consolidado, "w", encoding="latin1") as f:
                        f.write(texto)
                    print(f"  [OK] ANP {fuel} consolidado: {len(r.content)//1024} KB")
                    break
                except Exception as e:
                    print(f"  [AVISO] {url}: {e}")

        if texto:
            try:
                serie = _processar_anp_bruto(texto, col)
                # Filtra apenas dados a partir de 2020
                serie = serie[serie.index >= "2020-01-01"]
                resultado[col] = serie
                print(f"  [OK] ANP {fuel}: {len(serie)} pontos desde {serie.index.min().date()}")
            except Exception as e:
                print(f"  [AVISO] Falha ao processar ANP {fuel} consolidado: {e}")
        else:
            print(f"  [AVISO] ANP {fuel}: sem dados locais nem download disponível")

    if not resultado:
        print("  [AVISO] ANP: nenhum dado obtido — series glp_p13_brl e gnv_brl ausentes")
        return combinar_com

    df_anp = pd.DataFrame(resultado)
    df_anp.index.name = "date"
    df_anp.to_csv(os.path.join(RAW_DIR, "anp_consolidado.csv"))

    if combinar_com is not None and not combinar_com.empty:
        return combinar_com.join(df_anp, how="outer")
    return df_anp


def fetch_glp_campos(df_atual=None):
    """
    Lê data/raw/anp_glp_campos_YYYY.xlsx (planilha ANP com preços por campo: Tupi, Jubarte etc.)
    Calcula média ponderada mensal → coluna glp_campos_brl em prices.csv.
    Atualizar o arquivo anualmente (ou quando ANP publicar nova versão).
    """
    print("[3/6] GLP Campos — preços de gás natural por campo de produção...")
    ano_atual = date.today().year
    frames = []

    for ano in range(2020, ano_atual + 1):
        raw_path = os.path.join(RAW_DIR, f"anp_glp_campos_{ano}.xlsx")
        if not _arquivo_existe(raw_path):
            continue
        try:
            from openpyxl import load_workbook
            wb = load_workbook(raw_path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))

            header_idx = None
            peso_idx   = None
            for i, row in enumerate(all_rows):
                row_str = [str(v).strip() if v is not None else "" for v in row]
                row_lower = [x.lower() for x in row_str]
                if any(x in ["mês", "mes", "mÊs"] for x in row_str) or "mês" in row_lower:
                    header_idx = i
                if "peso" in row_lower:
                    peso_idx = i

            if header_idx is None:
                print(f"  [AVISO] {os.path.basename(raw_path)}: coluna Mês não encontrada")
                continue

            headers = [str(v).strip() if v is not None else f"col_{j}"
                       for j, v in enumerate(all_rows[header_idx])]
            mes_idx = next((j for j, h in enumerate(headers)
                            if h.lower() in ["mês", "mes"]), None)
            if mes_idx is None:
                continue

            pesos = {}
            if peso_idx is not None:
                for j, val in enumerate(all_rows[peso_idx]):
                    if j == mes_idx:
                        continue
                    h = headers[j] if j < len(headers) else None
                    if h and not h.startswith("col_") and val and isinstance(val, (int, float)):
                        pesos[h] = float(val)

            campo_cols = [h for j, h in enumerate(headers)
                          if j != mes_idx and h and not h.startswith("col_")
                          and not any(x in h.lower() for x in ["média","media","ponder","simpl"])]

            data_rows = []
            for row in all_rows[header_idx + 1:]:
                if not row or mes_idx >= len(row):
                    continue
                mes_val = row[mes_idx]
                if not mes_val:
                    continue
                try:
                    mes_str = str(mes_val).strip()
                    ts = pd.to_datetime(mes_str + "-01") if re.match(r"\d{4}-\d{2}", mes_str) else pd.to_datetime(mes_str, dayfirst=False, errors="coerce")
                    if pd.isna(ts):
                        continue
                except Exception:
                    continue
                row_dict = {"date": ts}
                for campo in campo_cols:
                    if campo in headers:
                        cidx = headers.index(campo)
                        v = row[cidx] if cidx < len(row) else None
                        if isinstance(v, (int, float)):
                            row_dict[campo] = float(v)
                data_rows.append(row_dict)

            if not data_rows:
                continue

            df = pd.DataFrame(data_rows).set_index("date")
            df = df[[c for c in df.columns if df[c].notna().any()]]
            cols_val = [c for c in df.columns if c in campo_cols]

            if pesos and any(c in pesos for c in cols_val):
                w = pd.Series({c: pesos.get(c, 1.0) for c in cols_val})
                w = w / w.sum()
                serie = df[cols_val].mul(w).sum(axis=1, min_count=1)
            else:
                serie = df[cols_val].mean(axis=1)

            serie.name = "glp_campos_brl"
            serie.index.name = "date"
            frames.append(serie.dropna())
            print(f"  [LOCAL] {os.path.basename(raw_path)}: {len(serie.dropna())} meses | campos: {cols_val[:4]}")

        except Exception as e:
            print(f"  [AVISO] GLP Campos {ano}: {e}")

    if not frames:
        print("  [INFO] Nenhum arquivo anp_glp_campos_YYYY.xlsx encontrado — coluna glp_campos_brl ignorada.")
        return df_atual

    serie_total = pd.concat(frames).sort_index()
    serie_total = serie_total[~serie_total.index.duplicated(keep="last")]
    serie_total.to_csv(os.path.join(RAW_DIR, "glp_campos_consolidado.csv"))
    print(f"  [OK] GLP Campos: {len(serie_total)} meses ({serie_total.index[0].strftime('%Y-%m')} → {serie_total.index[-1].strftime('%Y-%m')})")

    if df_atual is not None and not df_atual.empty:
        return df_atual.join(serie_total.to_frame(), how="outer")
    return serie_total.to_frame()


def fetch_ibge():
    print("[4/6] IBGE — PIB e população municipais...")
    fp_final = os.path.join(DATA_DIR, "ibge_mercado.csv")
    fp_raw   = os.path.join(RAW_DIR, "ibge_pib_bruto.csv")

    if _arquivo_existe(fp_raw):
        print("  [LOCAL] ibge_pib_bruto.csv encontrado em data/raw/ — processando...")
        try:
            df = pd.read_csv(fp_raw, low_memory=False)
            if "pib_total_mil" not in df.columns:
                raise ValueError(f"Coluna pib_total_mil não encontrada. Colunas: {list(df.columns)}")

            if "populacao" not in df.columns:
                print("  [INFO] Coluna populacao ausente — buscando da API IBGE...")
                try:
                    r = requests.get(
                        "https://servicodados.ibge.gov.br/api/v3/agregados/6579"
                        "/periodos/2022/variaveis/9324?localidades=N6[all]", timeout=60)
                    r.raise_for_status()
                    pop_rows = []
                    for item in r.json()[0]["resultados"][0]["series"]:
                        loc = item["localidade"]
                        val = list(item["serie"].values())[0]
                        if val and val not in ["-", None]:
                            nome  = loc["nome"]
                            match = re.search(r'\(([A-Z]{2})\)$', str(nome))
                            uf    = match.group(1) if match else None
                            mun   = re.sub(r'\s*\([A-Z]{2}\)$', '', str(nome)).strip()
                            pop_rows.append({"municipio": mun, "uf": uf, "populacao": int(val)})
                    df_pop = pd.DataFrame(pop_rows)
                    df = df.merge(df_pop, on=["municipio", "uf"], how="left")
                    print(f"  [OK] Populacao obtida da API: {df['populacao'].notna().sum()}/{len(df)} municipios")
                except Exception as e:
                    print(f"  [AVISO] API populacao nao acessivel: {e}")
                    df["populacao"] = None

            if "pib_per_capita" not in df.columns:
                df["pib_per_capita"] = (df["pib_total_mil"] * 1000 / df["populacao"]).round(0)
            if "uf" not in df.columns and "municipio" in df.columns:
                df["uf"] = df["municipio"].str.extract(r'\((\w{2})\)$')
                df["municipio"] = df["municipio"].str.replace(r'\s*\(\w{2}\)$', '', regex=True)
            df.to_csv(fp_final, index=False)
            print(f"  [OK] {len(df)} municipios processados de data/raw/ibge_pib_bruto.csv")
            return
        except Exception as e:
            print(f"  [ERRO] Falha ao processar ibge_pib_bruto.csv local: {e}")
            return

    if _cache_valido(fp_final, dias=28):
        print("  [CACHE] ibge_mercado.csv valido (< 28 dias) — pulando download")
        return

    print("  [DOWNLOAD] ibge_pib_bruto.csv nao encontrado, tentando API IBGE...")
    try:
        r = requests.get(
            "https://servicodados.ibge.gov.br/api/v3/agregados/5938"
            "/periodos/2023/variaveis/37?localidades=N6[all]", timeout=60)
        r.raise_for_status()
        rows = []
        for item in r.json()[0]["resultados"][0]["series"]:
            loc = item["localidade"]
            val = list(item["serie"].values())[0]
            if val and val not in ["-", None]:
                rows.append({"cod_municipio": loc["id"], "municipio": loc["nome"], "pib_total_mil": float(val)})
        df_pib = pd.DataFrame(rows)

        r2 = requests.get(
            "https://servicodados.ibge.gov.br/api/v3/agregados/6579"
            "/periodos/2022/variaveis/9324?localidades=N6[all]", timeout=60)
        r2.raise_for_status()
        pop_rows = []
        for item in r2.json()[0]["resultados"][0]["series"]:
            loc = item["localidade"]
            val = list(item["serie"].values())[0]
            if val and val not in ["-", None]:
                pop_rows.append({"cod_municipio": loc["id"], "populacao": int(val)})
        df_pop = pd.DataFrame(pop_rows)

        df = df_pib.merge(df_pop, on="cod_municipio", how="left")
        df["pib_per_capita"] = (df["pib_total_mil"] * 1000 / df["populacao"]).round(0)
        df["uf"] = df["municipio"].str.extract(r'\((\w{2})\)$')
        df["municipio"] = df["municipio"].str.replace(r'\s*\(\w{2}\)$', '', regex=True)
        df.to_csv(fp_raw, index=False)
        df.to_csv(fp_final, index=False)
        print(f"  [OK] {len(df)} municipios baixados da API IBGE")
    except Exception as e:
        print(f"  [AVISO] IBGE API: {e}")


def fetch_cbio():
    print("[5/6] CBIO — precos historicos...")
    fp_final = os.path.join(DATA_DIR, "cbio.csv")
    fp_raw   = os.path.join(RAW_DIR, "cbio_bruto.csv")

    # Arquivo local manual tem prioridade absoluta
    if _arquivo_existe(fp_raw):
        print("  [LOCAL] cbio_bruto.csv encontrado em data/raw/ — processando...")
        try:
            df = pd.read_csv(fp_raw, index_col=0, parse_dates=True)
            if "Close" in df.columns:
                df = df[["Close"]].rename(columns={"Close": "preco_cbio"})
            elif "preco_cbio" not in df.columns:
                raise ValueError(f"Coluna esperada nao encontrada. Colunas: {list(df.columns)}")
            else:
                df = df[["preco_cbio"]]
            df.index = pd.to_datetime(df.index).normalize()
            df.index.name = "date"
            df = df.dropna()
            df.index = df.index.strftime("%Y-%m-%d")
            df.to_csv(fp_final)
            print(f"  [OK] CBIO: {len(df)} semanas processadas")
            return
        except Exception as e:
            print(f"  [ERRO] cbio_bruto.csv: {e}")
            return

    # CBIO3.SA foi delistado — tenta tickers alternativos em ordem
    # CBIO11 é o ETF de CBIOs mais líquido negociado na B3
    CBIO_TICKERS = ["CBIO11.SA", "CBIO3.SA", "CBIO.SA"]
    print("  [DOWNLOAD] Tentando tickers CBIO no Yahoo Finance...")

    for ticker in CBIO_TICKERS:
        try:
            df = yf.download(ticker, period="3y", interval="1wk",
                             auto_adjust=True, progress=False)
            if df.empty:
                print(f"  [AVISO] {ticker}: sem dados")
                continue
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df = df[["Close"]].rename(columns={"Close": "preco_cbio"})
            df.index = pd.to_datetime(df.index).normalize()
            df.index.name = "date"
            df = df.dropna()
            if len(df) < 4:
                print(f"  [AVISO] {ticker}: dados insuficientes ({len(df)} pontos)")
                continue
            df.to_csv(fp_raw)
            df_save = df.copy()
            df_save.index = df_save.index.strftime("%Y-%m-%d")
            df_save.to_csv(fp_final)
            print(f"  [OK] CBIO via {ticker}: {len(df)} semanas | último: {df.index[-1].date()}")
            return
        except Exception as e:
            print(f"  [AVISO] {ticker}: {e}")

    print("  [INFO] CBIO: nenhum ticker disponível no Yahoo Finance.")
    print("         Para usar dados CBIO, coloque data/raw/cbio_bruto.csv")
    print("         (CSV com colunas: date, preco_cbio — exportado da B3 ou planilha própria)")


def salvar_prices(df):
    fp = os.path.join(DATA_DIR, "prices.csv")
    if os.path.exists(fp):
        existing = pd.read_csv(fp, index_col="date", parse_dates=True)
        # Novo dado tem prioridade; existente preenche apenas lacunas do novo
        df = df.combine_first(existing)
    df = df.sort_index()
    df.index = df.index.strftime("%Y-%m-%d")
    df.to_csv(fp)
    n_rows = len(df)
    n_cols = len(df.columns)
    n_ok   = df.notna().sum().sum()
    print(f"  [OK] prices.csv: {n_rows} linhas x {n_cols} series | {n_ok} valores nao-nulos | colunas: {list(df.columns)}")


def save_last_update():
    with open(os.path.join(DATA_DIR, "last_update.txt"), "w") as f:
        f.write(date.today().isoformat())
    print(f"  Ultima atualizacao: {date.today()}")


if __name__ == "__main__":
    print(f"\n{'='*60}")
    print(f"  CNP — Pipeline de dados  |  {date.today()}")
    print(f"{'='*60}\n")

    df_yahoo  = fetch_yahoo()
    df_prices = fetch_anp(combinar_com=df_yahoo if not df_yahoo.empty else None)
    if df_prices is None:
        df_prices = pd.DataFrame()
    df_prices = fetch_glp_campos(df_atual=df_prices)
    if df_prices is None:
        df_prices = pd.DataFrame()

    if not df_prices.empty:
        salvar_prices(df_prices)
    else:
        print("  [AVISO] Sem dados de precos — prices.csv nao atualizado.")

    fetch_ibge()
    fetch_cbio()
    save_last_update()

    print(f"\n{'='*60}")
    print("  Pipeline concluido!")
    print(f"{'='*60}\n")
